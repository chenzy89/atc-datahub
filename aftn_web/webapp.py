"""Flask Web 应用 — API + 前端页面"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from flask import Flask, jsonify, render_template, request, send_file

from .xlsx_writer import make_xlsx

from .config import AppConfig
from .database import Database
from .fdr_store import FDRStore
from .models import FlightPlan

logger = logging.getLogger("aftn_web.webapp")


from .radar_history import RadarHistoryStore
from .voice_receiver import VoiceReceiver

def create_app(config: AppConfig, db: Database, fdr_store: FDRStore | None = None, radar_history_store: RadarHistoryStore | None = None, voice_receiver: VoiceReceiver | None = None) -> Flask:
    _RADAR_MAP_VERSION = "v0.1"
    app = Flask(
        __name__,
        template_folder=str(Path(__file__).parent / "templates"),
    )

    # ── 页面 ──────────────────────────────────────────────────

    @app.route("/")
    def index():
        return render_template("index.html")

    @app.route("/api/maplist")
    def api_maplist():
        """返回 maplist.txt 中定义的地图文件列表"""
        map_dir = Path("/home/share/atc_aftn_web/map")
        lst = map_dir / "maplist.txt"
        files: list[str] = []
        if lst.exists():
            for line in lst.read_text("utf-8").splitlines():
                line = line.strip()
                if line and not line.startswith("#"):
                    files.append(line)
        return jsonify(files)

    @app.route("/api/map_file/<name>")
    def api_map_file(name: str):
        """读取 /home/share/atc_aftn_web/map/<name>.txt 地图文件"""
        import os
        safe = os.path.basename(name)
        filepath = Path("/home/share/atc_aftn_web/map") / f"{safe}.txt"
        if not filepath.exists():
            return jsonify({"error": "file not found"}), 404
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                with open(filepath, "r", encoding=enc) as f:
                    text = f.read()
                break
            except UnicodeDecodeError:
                continue
        else:
            return jsonify({"error": "decode failed"}), 500
        return text, 200, {"Content-Type": "text/plain; charset=utf-8"}

    # ── 雷达 FDR ──────────────────────────────────────────────

    @app.route("/api/radar_tracks")
    def api_radar_tracks():
        """返回当前 FDR 航迹列表（含位置）"""
        if fdr_store is None:
            return jsonify([])
        return jsonify(fdr_store.get_tracks())

    @app.route("/api/flight_info")
    def api_flight_info():
        """根据呼号+起降地查询完整飞行计划信息（雷达图弹出用）"""
        callsign = _req_str("callsign")
        adep = _req_str("adep")
        adest = _req_str("adest")
        if not callsign:
            return jsonify({"error": "callsign required"}), 400

        records = db.query_flight_plans(
            callsign=callsign,
            adep=adep,
            adest=adest,
            limit=5,
            offset=0,
        )
        if not records:
            # 降级：仅用呼号再查一次
            records = db.query_flight_plans(callsign=callsign, limit=5, offset=0)
        if not records:
            return jsonify({"error": "not found"}), 404

        # 选最新的（按 last_message_time）
        best = max(records, key=lambda r: r.get("last_message_time") or "")

        # 计算 EET
        eet_str = ""
        try:
            from datetime import datetime as _dt
            etd = best.get("etd")
            eta = best.get("eta")
            if etd and eta:
                if isinstance(etd, str): etd = _dt.fromisoformat(etd)
                if isinstance(eta, str): eta = _dt.fromisoformat(eta)
                minutes = int((eta - etd).total_seconds() / 60)
                eet_str = f"{minutes // 60:02d}:{minutes % 60:02d}"
        except Exception:
            pass

        return jsonify({
            "id": best["id"],
            "callsign": best.get("callsign", ""),
            "adep": best.get("adep", ""),
            "adest": best.get("adest", ""),
            "etd": _safe_dt(best.get("etd")),
            "atd": _safe_dt(best.get("atd")),
            "eta": _safe_dt(best.get("eta")),
            "ata": _safe_dt(best.get("ata")),
            "eet": eet_str,
            "aircraft_type": best.get("aircraft_type", ""),
            "route": best.get("route", ""),
            "handover_pt": best.get("handover_pt", ""),
            "runway": best.get("runway", ""),
            "flight_procedure": best.get("flight_procedure", ""),
            "entry_time": best.get("entry_time", ""),
            "exit_time": best.get("exit_time", ""),
            "terminal_flight_time": best.get("terminal_flight_time", 0),
            "ssr": best.get("ssr", ""),
        })

    # ── 雷达历史回放 ──────────────────────────────────────────

    @app.route("/api/radar_history")
    def api_radar_history():
        """返回时间范围内的历史航迹点"""
        if radar_history_store is None:
            return jsonify([])
        ts_from = _req_str("from")
        ts_to = _req_str("to")
        callsign = _req_str("callsign")
        if not ts_from or not ts_to:
            return jsonify({"error": "from and to required"}), 400
        return jsonify(radar_history_store.query(ts_from, ts_to, callsign))

    @app.route("/api/radar_time_range")
    def api_radar_time_range():
        """返回历史数据的时间范围"""
        if radar_history_store is None:
            return jsonify({"start": None, "end": None})
        start, end = radar_history_store.get_time_range()
        return jsonify({"start": start, "end": end})

    @app.route("/api/flight_trail")
    def api_flight_trail():
        """返回指定航班号当天的全部历史航迹点（用于飞行计划页的轨迹图）"""
        if radar_history_store is None:
            return jsonify([])
        callsign = _req_str("callsign")
        if not callsign:
            return jsonify({"error": "callsign required"}), 400
        date_str = _req_str("date") or datetime.utcnow().strftime("%Y-%m-%d")
        ts_from = f"{date_str}T00:00:00.000Z"
        ts_to = f"{date_str}T23:59:59.000Z"
        pts = radar_history_store.query(ts_from, ts_to, callsign)
        pts.sort(key=lambda p: p.get("ts", ""))
        return jsonify(pts)

    @app.route("/api/config/terminal_airports")
    def api_terminal_airports():
        """返回终端区完整配置（机场列表+多边形+高度）"""
        from .terminal_area import get_terminal_config_safe
        return jsonify(get_terminal_config_safe())

    @app.route("/api/fdr_stats")
    def api_fdr_stats():
        if fdr_store is None:
            return jsonify({"enabled": False, "total": 0, "associated": 0})
        stats = fdr_store.get_stats()
        return jsonify({"enabled": True, **stats})

    # ── 统计 ──────────────────────────────────────────────────

    @app.route("/api/statistics")
    def api_statistics():
        airports_str = _req_str("airports")
        date_from = _req_str("date_from")
        date_to = _req_str("date_to")
        airports = [a.strip().upper() for a in airports_str.split(",") if a.strip()] if airports_str else []
        if not airports or not date_from or not date_to:
            return jsonify({"error": "机场和日期范围必填"}), 400
        result = db.query_traffic_statistics(airports, date_from, date_to)
        result["sector_traffic"] = db.query_sector_traffic(date_from, date_to)
        return jsonify(result)

    @app.route("/api/statistics/export", methods=["POST"])
    def api_statistics_export():
        body = request.get_json()
        if not body:
            return jsonify({"error": "missing body"}), 400
        d = body.get("data", {})
        p = body.get("params", {})
        days = d.get("days", 1) or 1

        def fmt(sec):
            if not sec or sec <= 0:
                return "0m"
            h = sec // 3600
            m = (sec % 3600) // 60
            s = sec % 60
            return f"{h}h{m:02d}m{s:02d}s" if h else f"{m}m{s:02d}s"

        sheets = []

        # 概况
        summary = [
            ["指标", "值"],
            ["日期范围", f'{p.get("df","")} ~ {p.get("dt","")}'],
            ["机场", ", ".join(p.get("airports", []))],
            ["统计天数", days],
            ["日均出港", round(d.get("dep_count", 0) / days, 0)],
            ["日均进港", round(d.get("arr_count", 0) / days, 0)],
            ["出港总飞行时长", fmt(d.get("dep_terminal_seconds", 0))],
            ["进港总飞行时长", fmt(d.get("arr_terminal_seconds", 0))],
            ["高峰小时", f'{d.get("peak_hour", 0)}时({d.get("peak_hour_dep", 0) + d.get("peak_hour_arr", 0)}架次)'],
            ["高峰日", f'{d.get("peak_day", "")} {d.get("peak_day_count", 0)}架次'],
        ]
        sheets.append({"name": "概况", "rows": summary})

        # 进港移交点
        ah = d.get("arr_handover", {})
        arr_ho = [["移交点", "总架次", "日均", "占比"]]
        for k, v in sorted(ah.items(), key=lambda x: -x[1]):
            arr_ho.append([k, v, round(v / days, 0), f"{v / d.get('arr_count', 1) * 100:.1f}%"])
        sheets.append({"name": "进港移交点", "rows": arr_ho})

        # 出港移交点
        dh = d.get("dep_handover", {})
        dep_ho = [["移交点", "总架次", "日均", "占比"]]
        for k, v in sorted(dh.items(), key=lambda x: -x[1]):
            dep_ho.append([k, v, round(v / days, 0), f"{v / d.get('dep_count', 1) * 100:.1f}%"])
        sheets.append({"name": "出港移交点", "rows": dep_ho})

        # STAR 进港程序
        arr_term = d.get("arr_terminal_seconds", 0) or 0
        arr_proc = [["程序", "日均架次", "平均飞行时长", "时长占比"]]
        for proc in d.get("arr_procedures", []):
            avg = round(proc["total_seconds"] / proc["count"])
            pct = f"{proc['total_seconds'] / arr_term * 100:.1f}%" if arr_term else "-"
            arr_proc.append([proc["name"], round(proc["count"] / days, 1), fmt(avg), pct])
        sheets.append({"name": "进港程序STAR", "rows": arr_proc})

        # SID 出港程序
        dep_term = d.get("dep_terminal_seconds", 0) or 0
        dep_proc = [["程序", "日均架次", "平均飞行时长", "时长占比"]]
        for proc in d.get("dep_procedures", []):
            avg = round(proc["total_seconds"] / proc["count"])
            pct = f"{proc['total_seconds'] / dep_term * 100:.1f}%" if dep_term else "-"
            dep_proc.append([proc["name"], round(proc["count"] / days, 1), fmt(avg), pct])
        sheets.append({"name": "出港程序SID", "rows": dep_proc})

        # 24h流量
        dep_h = [round(v / days) for v in d.get("dep_hourly", [0] * 24)]
        arr_h = [round(v / days) for v in d.get("arr_hourly", [0] * 24)]
        hdr = ["小时"] + [f"{h}时" for h in range(24)] + ["日均"]
        hourly = [hdr]
        hourly.append(["进港"] + arr_h + [round(d.get("arr_count", 0) / days, 0)])
        hourly.append(["出港"] + dep_h + [round(d.get("dep_count", 0) / days, 0)])
        hourly.append(["合计"] + [arr_h[i] + dep_h[i] for i in range(24)] + [round((d.get("arr_count", 0) + d.get("dep_count", 0)) / days, 0)])
        sheets.append({"name": "小时流量", "rows": hourly})

        # 扇区流量
        st = d.get("sector_traffic", {})
        terminal_names = {
            "ZGJDTM01": "TM01 HN", "ZGJDTM02": "TM02 HE",
            "ZGJDTM03": "TM03 ARW", "ZGJDTM04": "TM04 AS",
            "ZGJDTM05": "TM05 AD", "ZGJDTM06": "TM06 ASL",
            "ZGJDTM07": "TM07 ARE/AA",
        }
        order = [f"ZGJDTM{i:02d}" for i in range(1, 8)]
        sec_hdr = ["扇区"] + [f"{h}时" for h in range(24)] + ["总计"]
        sector_rows = [sec_hdr]
        for code in order:
            data = st.get(code, [0]*24)
            daily = [round(v / days, 1) for v in data]
            total = round(sum(data) / days, 1)
            label = terminal_names.get(code, code)
            sector_rows.append([label] + daily + [total])
        sheets.append({"name": "扇区流量", "rows": sector_rows})

        xlsx_buf = make_xlsx(sheets)
        return send_file(
            xlsx_buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"统计_{p.get('df','')}_{p.get('dt','')}.xlsx",
        )

    @app.route("/api/stats")
    def api_stats():
        total_fpl = db.count_flight_plans()
        fpl_by_type: dict[str, int] = {}
        for t in ("FPL", "DEP", "ARR", "DLA"):
            fpl_by_type[t] = db.count_flight_plans(source_message_type=t)
        total_aftn = db.count_aftn_messages()
        aftn_by_type = db.count_aftn_by_type()
        return jsonify({
            "total_flight_plans": total_fpl,
            "flight_plans_by_type": fpl_by_type,
            "total_aftn_messages": total_aftn,
            "aftn_messages_by_type": aftn_by_type,
            "db_path": str(config.db_path),
        })

    # ══════════════════════════════════════════════════════════
    # 模块一：AFTN 报文
    # ══════════════════════════════════════════════════════════

    @app.route("/api/aftn_messages")
    def api_aftn_messages():
        msg_type = _req_str("message_type")
        keyword = _req_str("keyword")
        date_from = _req_str("date_from")
        date_to = _req_str("date_to")
        limit = request.args.get("limit", 100, type=int)
        offset = request.args.get("offset", 0, type=int)

        records = db.query_aftn_messages(
            message_type=msg_type,
            keyword=keyword,
            date_from=date_from,
            date_to=date_to,
            limit=min(limit, 500),
            offset=offset,
        )
        for rec in records:
            rec["sender_address"] = _extract_sender(rec.get("raw_text", ""))
        total = db.count_aftn_messages(
            message_type=msg_type,
            keyword=keyword,
            date_from=date_from,
            date_to=date_to,
        )
        return jsonify({"total": total, "records": records})

    # ══════════════════════════════════════════════════════════
    # 模块二：飞行计划
    # ══════════════════════════════════════════════════════════

    @app.route("/api/aftn_messages/<int:msg_id>")
    def api_aftn_message_get(msg_id: int):
        """获取单条 AFTN 报文"""
        record = db.get_aftn_message(msg_id)
        if record is None:
            return jsonify({"error": "not found"}), 404
        record["sender_address"] = _extract_sender(record.get("raw_text", ""))
        return jsonify(record)

    @app.route("/api/flight_plans")
    def api_flight_plans():
        """查询飞行计划列表（支持过滤+分页）"""
        callsign = _req_str("callsign")
        adep = _req_str("adep")
        adest = _req_str("adest")
        dof = _req_date("dof")
        airport = _req_str("airport")  # 关注机场：adep OR adest 匹配
        route = _req_str("route")  # 航路关键词
        source_message_type = _req_str("source_message_type")
        flight_rule = _req_str("flight_rule")
        handover_pt = _req_str("handover_pt")
        sort_by = _req_str("sort_by")  # callsign / etd
        sort_order = _req_str("sort_order")  # asc / desc
        limit = request.args.get("limit", 100, type=int)
        offset = request.args.get("offset", 0, type=int)

        records = db.query_flight_plans(
            callsign=callsign,
            adep=adep,
            adest=adest,
            dof=dof,
            airport=airport,
            route=route,
            source_message_type=source_message_type,
            flight_rule=flight_rule,
            handover_pt=handover_pt,
            sort_by=sort_by,
            sort_order=sort_order,
            limit=min(limit, 500),
            offset=offset,
        )
        total = db.count_flight_plans(
            callsign=callsign,
            adep=adep,
            adest=adest,
            dof=dof,
            airport=airport,
            route=route,
            source_message_type=source_message_type,
            flight_rule=flight_rule,
            handover_pt=handover_pt,
        )
        return jsonify({"total": total, "records": records})

    @app.route("/api/flight_plans/<int:fpl_id>")
    def api_flight_plan_get(fpl_id: int):
        """获取单条飞行计划"""
        record = db.get_flight_plan(fpl_id)
        if record is None:
            return jsonify({"error": "not found"}), 404
        return jsonify(record)

    @app.route("/api/flight_plans", methods=["POST"])
    def api_flight_plan_create():
        """手工新增飞行计划"""
        data = request.get_json(silent=True) or {}
        plan = _parse_plan_from_json(data)
        plan.id = 0
        try:
            new_id = db.create_flight_plan(plan)
            return jsonify({"id": new_id, "ok": True}), 201
        except Exception as exc:
            logger.exception("create flight plan failed")
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/flight_plans/<int:fpl_id>", methods=["PUT"])
    def api_flight_plan_update(fpl_id: int):
        """手工更新飞行计划"""
        data = request.get_json(silent=True) or {}
        plan = _parse_plan_from_json(data)
        plan.id = fpl_id
        ok = db.update_flight_plan(fpl_id, plan)
        if not ok:
            return jsonify({"error": "not found"}), 404
        return jsonify({"ok": True})

    @app.route("/api/flight_plans/<int:fpl_id>", methods=["DELETE"])
    def api_flight_plan_delete(fpl_id: int):
        """删除飞行计划"""
        ok = db.delete_flight_plan(fpl_id)
        if not ok:
            return jsonify({"error": "not found"}), 404
        return jsonify({"ok": True})

    @app.route("/api/flight_plans/export")
    def api_flight_plan_export():
        """导出飞行计划为 Excel (.xlsx)"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        callsign = _req_str("callsign")
        adep = _req_str("adep")
        adest = _req_str("adest")
        dof = _req_date("dof")
        airport = _req_str("airport")
        route = _req_str("route")
        source_message_type = _req_str("source_message_type")
        flight_rule = _req_str("flight_rule")
        handover_pt = _req_str("handover_pt")

        records = db.query_flight_plans(
            callsign=callsign,
            adep=adep,
            adest=adest,
            dof=dof,
            airport=airport,
            route=route,
            source_message_type=source_message_type,
            flight_rule=flight_rule,
            handover_pt=handover_pt,
            limit=10000,
            offset=0,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "飞行计划"

        headers = ["ID", "航班号", "规则与种类", "移交点", "机型", "DOF",
                   "起飞地", "ETD", "ATD", "目的地", "ETA", "ATA",
                   "跑道", "飞行程序", "航路", "报文", "报文时间"]
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, rec in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=rec.get("id", ""))
            ws.cell(row=row_idx, column=2, value=rec.get("callsign", ""))
            ws.cell(row=row_idx, column=3, value=rec.get("flight_rule", ""))
            ws.cell(row=row_idx, column=4, value=rec.get("handover_pt", ""))
            ws.cell(row=row_idx, column=5, value=rec.get("aircraft_type", ""))
            ws.cell(row=row_idx, column=6, value=_safe_date(rec.get("dof")))
            ws.cell(row=row_idx, column=7, value=rec.get("adep", ""))
            ws.cell(row=row_idx, column=8, value=_safe_dt(rec.get("etd")))
            ws.cell(row=row_idx, column=9, value=_safe_dt(rec.get("atd")))
            ws.cell(row=row_idx, column=10, value=rec.get("adest", ""))
            ws.cell(row=row_idx, column=11, value=_safe_dt(rec.get("eta")))
            ws.cell(row=row_idx, column=12, value=_safe_dt(rec.get("ata")))
            ws.cell(row=row_idx, column=13, value=rec.get("runway", ""))
            ws.cell(row=row_idx, column=14, value=rec.get("flight_procedure", ""))
            ws.cell(row=row_idx, column=15, value=rec.get("route", ""))
            ws.cell(row=row_idx, column=16, value=rec.get("message_types", ""))
            ws.cell(row=row_idx, column=17, value=_safe_dt(rec.get("last_message_time")))
            for col in range(1, 18):
                ws.cell(row=row_idx, column=col).border = thin_border
                ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center")

        col_widths = [6, 14, 10, 12, 12, 10, 8, 12, 8, 16, 16, 10, 16, 16, 50, 10, 30, 16]
        for i, w in enumerate(col_widths, 1):
            # column letter (A-Z, AA, AB...)
            letter = ""
            n = i
            while n > 0:
                n -= 1
                letter = chr(65 + n % 26) + letter
                n //= 26
            ws.column_dimensions[letter].width = w

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue(), 200, {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f"attachment; filename=flight_plans_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
        }


    # ── 辅助 ──

    def _safe_date(v: Any) -> str:
        return str(v)[:10] if v else ""

    def _safe_dt(v: Any) -> str:
        return str(v)[:16] if v else ""

    def _extract_sender(raw_text: str) -> str:
        """从 AFTN 报文中提取发报地址"""
        if not raw_text:
            return ""
        parts = raw_text.split()
        for i, p in enumerate(parts):
            if p in ("FF", "GG") and i + 3 < len(parts):
                cand = parts[i + 3].upper()
                if len(cand) in (7, 8) and cand.isalpha():
                    return cand
        return ""

    # ── 辅助 ──────────────────────────────────────────────────

    def _req_str(key: str) -> Optional[str]:
        v = request.args.get(key, "").strip()
        return v if v else None

    def _req_date(key: str) -> Optional[date]:
        v = request.args.get(key, "").strip()
        if not v:
            return None
        try:
            return date.fromisoformat(v)
        except ValueError:
            return None

    def _parse_plan_from_json(data: dict[str, Any]) -> FlightPlan:
        def _dt(v: Any) -> datetime | None:
            if not v:
                return None
            if isinstance(v, datetime):
                return v
            for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(str(v)[:19], fmt)
                except ValueError:
                    continue
            return None

        def _d(v: Any) -> date | None:
            if not v:
                return None
            if isinstance(v, date):
                return v
            for fmt in ("%Y-%m-%d",):
                try:
                    return datetime.strptime(str(v)[:10], fmt).date()
                except ValueError:
                    continue
            return None

        return FlightPlan(
            id=int(data.get("id") or 0),
            callsign=str(data.get("callsign") or "").strip().upper(),
            ssr=str(data.get("ssr") or "").strip().upper(),
            aircraft_type=str(data.get("aircraft_type") or "").strip().upper(),
            dof=_d(data.get("dof")),
            adep=str(data.get("adep") or "").strip().upper(),
            etd=_dt(data.get("etd")),
            atd=_dt(data.get("atd")),
            adest=str(data.get("adest") or "").strip().upper(),
            eta=_dt(data.get("eta")),
            ata=_dt(data.get("ata")),
            route=str(data.get("route") or "").strip().upper(),
            source_message_type=str(data.get("source_message_type") or "MANUAL"),
            last_message_time=_dt(data.get("last_message_time")),
            raw_message_text=str(data.get("raw_message_text") or ""),
        )

    # ── 语音数据 ──────────────────────────────────────────

    @app.route("/api/voice/status")
    def api_voice_status():
        """返回语音通道状态"""
        if voice_receiver is None:
            return jsonify({"enabled": False, "channels": []})
        return jsonify({
            "enabled": True,
            "playing": voice_receiver.get_playing_channel(),
            "channels": voice_receiver.get_status(),
        })

    @app.route("/api/voice/select", methods=["POST"])
    def api_voice_select():
        """选择播放通道"""
        if voice_receiver is None:
            return jsonify({"error": "voice not enabled"}), 400
        body = request.get_json(silent=True) or {}
        channel = body.get("channel", -1)
        ok = voice_receiver.select_channel(channel)
        if not ok:
            return jsonify({"error": "invalid channel"}), 400
        return jsonify({"ok": True, "channel": channel})

    _voice_config = {
        "retention_days": config.voice_data.retention_days,
        "flight_count_max": config.voice_data.flight_count_max,
        "save_dir": str(config.voice_data.save_dir or ""),
    } if hasattr(config, "voice_data") else {
        "retention_days": 30,
        "flight_count_max": 18,
        "save_dir": "",
    }

    @app.route("/api/voice/config")
    def api_voice_config():
        """返回语音配置参数"""
        return jsonify(_voice_config)

    @app.route("/api/voice/duration")
    def api_voice_duration():
        """返回指定通道在指定日期的 144 个通话时长 + 扇区飞行架次"""
        if voice_receiver is None:
            return jsonify({"error": "voice not enabled"}), 400
        date_str = request.args.get("date", datetime.utcnow().strftime("%Y-%m-%d"))
        channel_str = request.args.get("channel", "38")
        try:
            channel_id = int(channel_str)
        except (ValueError, TypeError):
            return jsonify({"error": "invalid channel"}), 400

        # 通道号 → 终端扇区代码
        from .voice_receiver import CHANNEL_SECTORS
        terminal_code = CHANNEL_SECTORS.get(channel_id, "")

        # 语音时长
        duration_data = voice_receiver.get_channel_duration(date_str, channel_id)

        # 扇区飞行架次（10 分钟粒度）
        flight_data = []
        if terminal_code:
            try:
                flight_data = db.query_sector_traffic_10min(date_str, terminal_code)
            except Exception:
                pass

        return jsonify({
            "date": date_str,
            "channel": channel_id,
            "terminal_code": terminal_code,
            "duration": duration_data,
            "flight_count": flight_data,
            "flight_count_max": _voice_config["flight_count_max"],
        })

    @app.route("/api/voice/recordings")
    def api_voice_recordings():
        """返回指定日期和通道的录音文件列表"""
        if voice_receiver is None:
            return jsonify({"error": "voice not enabled"}), 400
        date_str = request.args.get("date", "")
        channel_str = request.args.get("channel", "")
        if not date_str or not channel_str:
            return jsonify({"error": "date and channel required"}), 400
        try:
            channel_id = int(channel_str)
        except (ValueError, TypeError):
            return jsonify({"error": "invalid channel"}), 400

        recordings = voice_receiver.list_recordings(date_str, channel_id)
        total_size = sum(r["size"] for r in recordings)
        return jsonify({
            "date": date_str,
            "channel": channel_id,
            "recordings": recordings,
            "count": len(recordings),
            "total_size": total_size,
            "total_size_str": voice_receiver._format_size(total_size),
        })

    @app.route("/api/voice/play_file")
    def api_voice_play_file():
        """返回指定日期/通道/时间范围的录音，解码为 WAV 流"""
        if voice_receiver is None:
            return jsonify({"error": "voice not enabled"}), 400
        date_str = request.args.get("date", "")
        channel_str = request.args.get("channel", "")
        from_time = request.args.get("from", "")
        to_time = request.args.get("to", "")
        if not date_str or not channel_str:
            return jsonify({"error": "date and channel required"}), 400
        try:
            channel_id = int(channel_str)
        except (ValueError, TypeError):
            return jsonify({"error": "invalid channel"}), 400

        wav_data = voice_receiver.get_recording_data(
            date_str, channel_id, from_time=from_time, to_time=to_time
        )
        if not wav_data:
            return jsonify({"error": "no recordings found"}), 404

        from io import BytesIO
        buf = BytesIO(wav_data)
        return send_file(
            buf,
            mimetype="audio/wav",
            as_attachment=False,
            download_name=f"voice_{date_str}_ch{channel_id}.wav",
        )

    @app.route("/api/voice/dates")
    def api_voice_dates():
        """返回指定通道下有录音的日期列表"""
        if voice_receiver is None:
            return jsonify({"error": "voice not enabled"}), 400
        channel_str = request.args.get("channel", "")
        try:
            channel_id = int(channel_str)
        except (ValueError, TypeError):
            return jsonify({"error": "invalid channel"}), 400
        dates = voice_receiver.list_dates(channel_id)
        return jsonify({"dates": dates})

    @app.route("/api/voice/stop", methods=["POST"])
    def api_voice_stop():
        """停止播放"""
        if voice_receiver is None:
            return jsonify({"error": "voice not enabled"}), 400
        voice_receiver.select_channel(-1)
        return jsonify({"ok": True})

    # ── 禁用浏览器缓存 API 响应 ──────────────────────────
    @app.after_request
    def _no_cache(response):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

    return app
